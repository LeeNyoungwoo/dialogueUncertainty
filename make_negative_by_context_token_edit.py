import argparse
import os
import pickle

# pip install uncertainty-calibration
import calibration as cal
import numpy as np
import scipy
import tensorflow_hub as hub
import torch
import torch.nn as nn
import transformers
from matplotlib import pyplot as plt
from sklearn.metrics import accuracy_score, f1_score
from sklearn.metrics.pairwise import cosine_similarity
from tensorboardX import SummaryWriter
from openpyxl.styles.colors import Color
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from torch import Tensor
from torch.nn import functional as F
from torch.nn.modules.loss import CrossEntropyLoss
from torch.optim.adamw import AdamW
from torch.utils.data import DataLoader, Dataset, RandomSampler
from torch.utils.data.dataloader import DataLoader
from tqdm import tqdm
from transformers import (
    BertConfig,
    BertForMaskedLM,
    BertForNextSentencePrediction,
    BertModel,
    BertTokenizer,
)

from preprocess_dataset import get_dd_corpus
from selection_model import BertSelect
from utils import (
    RankerDataset,
    get_nota_token,
    get_uttr_token,
    load_model,
    set_random_seed,
    write2tensorboard,
)
from string import ascii_uppercase

ascii_uppercase = list(ascii_uppercase)
column_titles = (
    ascii_uppercase
    + ["A" + c for c in ascii_uppercase]
    + ["B" + c for c in ascii_uppercase]
    + ["C" + c for c in ascii_uppercase]
    + ["D" + c for c in ascii_uppercase]
    + ["E" + c for c in ascii_uppercase]
)


def main(args):
    set_random_seed(42)

    device = torch.device("cuda")
    tokenizer = BertTokenizer.from_pretrained("bert-base-uncased")
    UTTR_TOKEN = get_uttr_token()
    NOTA_TOKEN = get_nota_token()
    special_tokens_dict = {"additional_special_tokens": [UTTR_TOKEN, NOTA_TOKEN]}
    tokenizer.add_special_tokens(special_tokens_dict)

    bert = BertModel.from_pretrained("bert-base-uncased")
    bert.resize_token_embeddings(len(tokenizer))
    model = BertSelect(bert)
    model = load_model(model, "./logs/select_batch12_candi8/model", 0, len(tokenizer))
    model.to(device)

    """
    STEP1. Get attention map and the orignal prediction score
    """
    prediction_list = []
    print("STEP1: Draw attention map")
    for setname in ["test", "dev", "train"]:
        output_fname = args.attention_dump_fname.format(setname)
        if os.path.exists(output_fname):
            continue

        raw_dd_dataset = get_dd_corpus(setname if setname != "dev" else "validation")
        mydataset = RankerDataset(raw_dd_dataset, tokenizer, setname, 300, UTTR_TOKEN)
        saver = []
        for idx in tqdm(range(len(mydataset))):
            with torch.no_grad():
                sample = [el[idx] for el in mydataset.feature]
                ids, masks = [torch.unsqueeze(el, 0).to(device) for el in sample[:2]]
                # 1 for positive and 0 for random
                if int(sample[2].numpy()) != 1:
                    continue
                prediction, attentions = model.get_attention(ids, masks)
                attention_output = [el.cpu().numpy() for el in attentions]
                attention_output = sum([sum(sum(el[0])) for el in attention_output])
                attention_output = [el for el in attention_output if el != 0]
                prediction = float(prediction.cpu().numpy())
                assert isinstance(prediction, float)
                prediction_list.append(prediction)

            token_list = tokenizer.convert_ids_to_tokens(
                [int(el) for el in mydataset.feature[0][idx].numpy() if el != 0]
            )
            assert len(token_list) == len(attention_output)
            for tok_idx, tok in enumerate(token_list):
                if tok in [".", ",", "?", "!", "[UTTR]", "[CLS]", "[SEP]"]:
                    attention_output[tok_idx] = 0

            final_item = {
                "prediction": prediction,
                "attention": attention_output,
                "feature": [int(el) for el in mydataset.feature[0][idx].numpy() if el != 0],
                "tokens": token_list,
            }

            saver.append(final_item)

        with open(output_fname, "wb") as f:
            pickle.dump(saver, f)
        continue
    """
    STEP2: choose the tokens to be masked & Replace them
    """
    print("STEP2")
    for setname in ["dev", "train"]:
        input_fname = args.attention_dump_fname.format(setname)
        assert os.path.exists(input_fname)
        changed_output_fname = args.changed_context_dump_fname.format(args.change_ratio, setname)
        if os.path.exists(changed_output_fname):
            print(changed_output_fname, " exist!")
            continue
        else:
            print(changed_output_fname, "not exist!")
        with open(input_fname, "rb") as f:
            attention_data = pickle.load(f)
        raw_dd_dataset = get_dd_corpus(setname if setname != "dev" else "validation")
        mydataset = RankerDataset(raw_dd_dataset, tokenizer, setname, 300, UTTR_TOKEN)
        changed_context_data = edit_context(
            args, attention_data, tokenizer, model, device, use_mlm=False, use_unk=True
        )
        with open(changed_output_fname, "wb") as f:
            pickle.dump(changed_context_data, f)

    """
    STEP3: Save results into excel
    """
    '''
    print("STEP3")
    for setname in ["dev", "train"]:
        changed_output_fname = args.changed_context_dump_fname.format(setname)
        with open(changed_output_fname, "rb") as f:
            data = pickle.load(f)
        data = [
            el
            for el in data
            if 0.3 < el["use_cosine_similarity"] < 0.8
            and el["changed_prediction"] - el["prediction"] < -0.2
        ]

        wb = Workbook()
        ws = wb.active
        for item_idx, item in enumerate(data):
            if item_idx == 1000:
                break
            original_sequence = tokenizer.tokenize(item["original_total_sequence"])
            changed_context = tokenizer.tokenize(item["changed_context"])
            response = original_sequence[len(changed_context) :][:-1]
            original_context = original_sequence[1 : len(changed_context) - 1]
            changed_context = changed_context[1:-1]
            assert len(changed_context) == len(original_context)
            """
            Index
            Original Context
            Changed Context
            Response
            Original Pred.
            Changed Pred.
            USE similarity
            \n (space)
            """

            diff_idx = [
                idx
                for idx in range(len(original_context))
                if original_context[idx] != changed_context[idx]
            ]
            ws["A" + str(item_idx * 8 + 1)] = "Index"
            ws["A" + str(item_idx * 8 + 1)].font = Font(bold=True)
            ws["B" + str(item_idx * 8 + 1)] = item_idx
            ws["A" + str(item_idx * 8 + 2)] = "Original Context"
            ws["A" + str(item_idx * 8 + 2)].font = Font(bold=True)
            for idx, tok in enumerate(original_context):
                ws[column_titles[1 + idx] + str(item_idx * 8 + 2)] = tok
                if idx in diff_idx:
                    ws[column_titles[1 + idx] + str(item_idx * 8 + 2)].font = Font(color="FF0000")
            ws["A" + str(item_idx * 8 + 3)] = "Channged Context"
            ws["A" + str(item_idx * 8 + 3)].font = Font(bold=True)
            for idx, tok in enumerate(changed_context):
                ws[column_titles[1 + idx] + str(item_idx * 8 + 3)] = tok
                if idx in diff_idx:
                    ws[column_titles[1 + idx] + str(item_idx * 8 + 3)].font = Font(color="FF0000")
            ws["A" + str(item_idx * 8 + 4)] = "Response"
            ws["A" + str(item_idx * 8 + 4)].font = Font(bold=True)
            for idx, tok in enumerate(response):
                ws[column_titles[1 + idx] + str(item_idx * 8 + 4)] = tok
            ws["A" + str(item_idx * 8 + 5)] = "Original Pred."
            ws["B" + str(item_idx * 8 + 5)] = round(item["prediction"], 2)
            ws["A" + str(item_idx * 8 + 6)] = "Changed Pred."
            ws["B" + str(item_idx * 8 + 6)] = round(item["changed_prediction"], 2)
            ws["A" + str(item_idx * 8 + 7)] = "Context Similarity"
            ws["B" + str(item_idx * 8 + 7)] = round(item["use_cosine_similarity"], 2)
        wb.save("changed_context_analysis.xlsx")
    '''


def edit_context(
    args, attention_dataset, tokenizer: BertTokenizer, model, device, use_mlm, use_unk
):
    """
    Heuristic하게 context를 바꿔보자!
    use_mlm: MLM으로 빈 칸을 다시 채울지 말지
    use_unk: MLM으로 빈 칸을 안채우고 그냥 UNK로 두기
    """
    assert sum([use_mlm, use_unk]) == 1
    if use_mlm:
        use_model = hub.load("https://tfhub.dev/google/universal-sentence-encoder/4")

    softmax_dim1 = torch.nn.Softmax(dim=1)
    tokens_to_skip = tokenizer.convert_tokens_to_ids(
        [".", ",", "?", "!", "[UTTR]", "[CLS]", "[SEP]"]
    )

    saver = []
    statistic_counter = {
        "original_low": 0,
        "success": 0,
        "original_similarity": 0,
        "score_no_down": 0,
        "short_context": 0,
    }
    bert = BertForMaskedLM.from_pretrained("bert-base-uncased")
    bert.to(device)

    uttr_token_id = tokenizer.convert_tokens_to_ids(["[UTTR]"])[0]
    assert isinstance(uttr_token_id, int)
    for item_idx, item in enumerate(tqdm(attention_dataset)):
        original_prediction, attention, input_ids = (
            item["prediction"],
            item["attention"],
            item["feature"],
        )

        assert len(attention) == len(input_ids)
        assert input_ids.count(tokenizer.sep_token_id) == 2

        if original_prediction < args.original_prediction_threshold:
            statistic_counter["original_low"] += 1
            continue

        context = input_ids[: input_ids.index(tokenizer.sep_token_id) + 1][:]

        attention = attention[: len(context)]
        for idx, tok in enumerate(context):
            if tok in tokens_to_skip:
                attention[idx] = 0
        attention /= sum(attention)
        if len(attention) - np.count_nonzero(attention) < args.context_minimum_length:
            statistic_counter["short_context"] += 1
            continue

        selected_tok_id_in_context = np.argsort(attention)[::-1][
            : int(len(attention) * args.change_ratio)
        ]
        original_token_list = [
            tok_id
            for tok_idx, tok_id in enumerate(context)
            if tok_idx in selected_tok_id_in_context
        ]

        if use_mlm:
            masked_context = [
                tokenizer.mask_token_id if tok_idx in selected_tok_id_in_context else tok_id
                for tok_idx, tok_id in enumerate(context)
            ]
            uttr_index_list = np.where(np.array(masked_context) == uttr_token_id)[0]
            masked_context = [el for el in masked_context if el != uttr_token_id]
            masked_token_index = [
                idx
                for idx in range(len(masked_context))
                if masked_context[idx] == tokenizer.mask_token_id
            ]

            with torch.no_grad():
                output = bert(torch.tensor([masked_context]).to(device))[0]

            for mask_order, mask_index in enumerate(masked_token_index):
                output[0][mask_index, original_token_list[mask_order]] = -100
                decoded_index = torch.argmax(output[0][mask_index]).item()
                masked_context[mask_index] = decoded_index

            """
            Similarity 
            """
            original = tokenizer.decode(context[1:-1])
            changed = tokenizer.decode(masked_context[1:-1])
            original_emb = use_model([original])
            changed_emb = use_model([changed])
            assert len(original_emb) == len(changed_emb) == 1
            cossim = cosine_similarity(original_emb, changed_emb)
            assert len(cossim) == 1 and len(cossim[0]) == 1
            cossim = float(cossim[0][0])
            item["use_cosine_similarity"] = cossim
            """
            Ranker score with changed context
            """
            changed_context_with_uttr_token = masked_context[:]
            for idx in uttr_index_list:
                changed_context_with_uttr_token.insert(idx, uttr_token_id)
            changed_context_with_uttr_token_str = tokenizer.decode(changed_context_with_uttr_token)
            assert isinstance(changed_context_with_uttr_token_str, str)
            item["changed_context"] = changed_context_with_uttr_token_str
            item["changed_total_sequence"] = tokenizer.decode(
                changed_context_with_uttr_token
                + input_ids[input_ids.index(tokenizer.sep_token_id) + 1 :]
            )
            item["original_total_sequence"] = tokenizer.decode(input_ids)
            with torch.no_grad():
                output = model(
                    torch.tensor(
                        [
                            changed_context_with_uttr_token
                            + input_ids[input_ids.index(tokenizer.sep_token_id) + 1 :]
                        ]
                    ).to(device),
                    output_attentions=True,
                    return_dict=True,
                )
                prediction = softmax_dim1(output["logits"]).cpu().numpy()[0][1]
            item["changed_prediction"] = float(prediction)
        else:
            unked_context = [
                tokenizer.unk_token_id if tok_idx in selected_tok_id_in_context else tok_id
                for tok_idx, tok_id in enumerate(context)
            ]
            item["changed_context"] = unked_context
            item["changed_total_sequence"] = tokenizer.decode(unked_context) + " ".join(
                [
                    "[NOTA]",
                    tokenizer.sep_token,
                ]
            )

        saver.append(item)
    return saver


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Process some integers.")
    parser.add_argument("--log_path", type=str, default="logs")
    parser.add_argument("--epoch", type=int, default=1)
    parser.add_argument("--attention_dump_fname", type=str, default="./attention/attention_{}.pck")
    parser.add_argument("--attention_img_fname", type=str, default="img/attn_{}.png")
    parser.add_argument(
        "--changed_context_dump_fname", type=str, default="./attention/UW_attention_change{}_{}.pck"
    )
    parser.add_argument(
        "--scoring_method", type=str, default="attention", choices=["attention", "adversarial"]
    )

    """
    Hyperparameters for context-editing process
    """
    parser.add_argument(
        "--context_minimum_length",
        type=int,
        default=5,
        help="[CLS],[SEP],[UTTR]이나 .,?!같은 특수 토큰을 제외하고 attention score를 부여받은 단어의 최소 개수",
    )
    parser.add_argument(
        "--original_prediction_threshold",
        type=float,
        default=-10000,
        help="어떤 sample에 대한 original confidence score가 이거보다 낮으면 context를 바꿔서 새로운 sample을 만들지 않음",
    )
    parser.add_argument(
        "--change_ratio",
        type=float,
        default=0.2,
        help="attention score가 높은 애들 중에서 이만큼 골라서 바꿈",
    )
    parser.add_argument(
        "--similarity_threshold",
        type=float,
        default=0.8,
        help="원래 context랑 바뀐 context랑 이거 이상으로 유사하면 잘 안바뀐걸로 치고, 이걸 최종 결과에 포함하지 않음",
    )
    parser.add_argument(
        "--lowered_score_gap",
        type=float,
        default=0.3,
        help="바뀐 context랑 원래 response에 대해 score를 재 보았을 때 얼마나 줄어드는지.",
    )

    args = parser.parse_args()
    os.makedirs(os.path.dirname(args.attention_dump_fname), exist_ok=True)
    main(args)